from telebot import TeleBot, types
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string

wb = load_workbook('./prices.xlsx')
sheet = wb.get_sheet_by_name('Лист1')
column_index_from_string('A')
print(sheet['A25'].value)


bot = TeleBot('1960085788:AAFx6PskBD-PhWfmi2qUxxoG7BX29VqFOF4')

@bot.message_handler(commands=['price'])
def amd_command(message):

    photo = open('./price.jpg', 'rb')
    bot.send_photo(message.chat.id, photo)

bot.polling()

